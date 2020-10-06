from datetime import datetime
from openpyxl import load_workbook
from classes import Product,Review
from mapping import PRODUCT_ID, PRODUCT_HEALTH_CAMP, REVIEW_ID,\
                    REVIEW_VAR1, REVIEW_VAR2, REVIEW_VAR3
                    #REVIEW_DATE

workbook = load_workbook(filename="sample1.xlsx", read_only=True)
sheet=workbook.active

products=[]
reviews=[]


for row in sheet.iter_rows(min_row=2, values_only=True):
    product = Product(id = row[PRODUCT_ID], Health_camp= row[PRODUCT_HEALTH_CAMP])

    products.append(product)

    #spread_date = col[REVIEW_DATE]
    #parsed_date = datetime.strptime(spread_date, "%d-%m-%Y")

    review = Review(id = row[REVIEW_ID], Var1 = row[REVIEW_VAR1], Var2 = row[REVIEW_VAR2], Var3 = row[REVIEW_VAR3])
                    #,date = parsed_date)

    reviews.append(review)


print(products[0])
print(reviews[0])
